import time
from typing import Generator, List

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from selenium import webdriver
from selenium.webdriver.remote.webdriver import WebDriver
from selenium.webdriver.common.by import By
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.common.action_chains import ActionChains

import config
from Person import Person
from config import target_url


def import_check_data(username: str, password: str, file_path: str, mode: int):
    switcher = {
        0: import_new,
        1: import_cover
    }
    importer = switcher.get(mode)
    importer(username, password, file_path)


def import_new(username: str, password: str, file_path: str):
    driver: WebDriver = webdriver.Chrome()
    driver.maximize_window()
    driver.get(target_url)
    set_username(driver, username)
    set_password(driver, password)
    select_role(driver)
    click_login_button(driver)
    click_health_doc_li(driver)
    click_health_check_sheet_li(driver)
    click_limit_select_img(driver)
    click_id_limit_select(driver)
    persons: List[Person] = read_excel(file_path)
    for person in persons:
        enter_person_id(driver, person.id)
        click_search_button(driver)
        double_click_result_list(driver)
        click_create_button(driver)
        import_person_data(driver, person)
        config.current_row_index += 1


def import_person_data(driver: WebDriver, person: Person):
    import_check_way(driver, person)
    import_symptom(driver, person)
    import_common(driver, person)
    import_life_style(driver, person)


def import_life_style(driver: WebDriver, person: Person):
    click_select_if_all_no_select(driver, "physicalExerciseFrequency", "4")
    click_select_if_all_no_select(driver, "dietaryHabit", "1")
    click_select_if_all_no_select(driver, "wehtherSmoke", "1")
    click_select_if_all_no_select(driver, "drinkingFrequencydrinkingFrequency", "1")
    click_select_if_all_no_select(driver, "occupational", "1")
    click_select_if_all_no_select(driver, "lip", "1")
    import_denture(driver, person)
    click_select_if_all_no_select(driver, "lip", "1")
    click_select_if_all_no_select(driver, "pharyngeal", "1")


def import_eye_sight(driver: WebDriver, person: Person):
    send_keys_by_xpath(driver, path="//input[@name='leftEye']", keys=person.left_eye)
    send_keys_by_xpath(driver, path="//input[@name='rightEye']", keys=person.right_eye)


def import_denture(driver: WebDriver, person: Person):
    if not person.is_elder():
        click_select_if_all_no_select(driver, "denture", "1")
        return
    click_select_by_xpath(driver, path="//input[@name='denture'][@value=2]")
    send_keys_by_xpath(driver, path="//input[@name='leftUp']", keys=person.teeth_left_up)
    send_keys_by_xpath(driver, path="//input[@name='leftDown']", keys=person.teeth_left_down)
    send_keys_by_xpath(driver, path="//input[@name='rightUp']", keys=person.teeth_right_up)
    send_keys_by_xpath(driver, path="//input[@name='rightDown']", keys=person.teeth_right_down)


def click_select_if_all_no_select(driver: WebDriver, name: str, value: str):
    elements_path = f"//input[@name={name}]"
    elements: List[WebElement] = find_elements_by_path(driver, elements_path)
    if not is_all_no_select(elements):
        return
    target_path = f"//input[@name={name}][@value={value}]"
    click_select_by_xpath(driver, target_path)


def is_all_no_select(elements: List[WebElement]):
    for element in elements:
        if element.is_selected():
            return False
    return True


def find_elements_by_path(driver: WebDriver, path: str) -> List[WebElement]:
    return driver.find_elements(By.XPATH, path)


def import_common(driver: WebDriver, person: Person):
    enter_temperature(driver, person)
    enter_pulse_rate(driver, person)
    enter_right_systolic_pressure(driver, person)
    enter_blood_pressure(driver, person)
    enter_height(driver, person)
    enter_weight(driver, person)
    enter_waistline(driver, person)
    click_bmi_input(driver)
    click_elder_rel_select(driver, person)


def click_elder_rel_select(driver: WebDriver, person: Person):
    if not person.is_elder():
        return
    click_select_by_xpath(driver, path="//input[@name='healthStatus'][@value=1]")
    click_select_by_xpath(driver, path="//input[@name='selfCare'][@value=1]")
    click_select_by_xpath(driver, path="//input[@name='cognitive'][@value=1]")
    click_select_by_xpath(driver, path="//input[@name='emotion'][@value=1]")


def click_bmi_input(driver: WebDriver):
    find_element_by_xpath(driver, path="//input[@name='bmi']").click()


def enter_waistline(driver: WebDriver, person: Person):
    send_keys_by_xpath(driver, path="//input[@name='waistline']", keys=person.waistline)


def enter_weight(driver: WebDriver, person: Person):
    send_keys_by_xpath(driver, path="//input[@name='weight']", keys=person.weight)


def enter_height(driver: WebDriver, person: Person):
    send_keys_by_xpath(driver, path="//input[@name='height']", keys=person.height)


def enter_blood_pressure(driver: WebDriver, person: Person):
    enter_right_systolic_pressure(driver, person)
    enter_right_diastolic_pressure(driver, person)
    enter_left_systolic_pressure(driver, person)
    enter_left_diastolic_pressure(driver, person)


def enter_left_diastolic_pressure(driver: WebDriver, person: Person):
    send_keys_by_xpath(driver, path="//input[@name='diastolic_L']", keys=person.left_diastolic_pressure)


def enter_left_systolic_pressure(driver: WebDriver, person: Person):
    send_keys_by_xpath(driver, path="//input[@name='constriction_L']", keys=person.left_systolic_pressure)


def enter_right_diastolic_pressure(driver: WebDriver, person: Person):
    send_keys_by_xpath(driver, path="//input[@name='diastolic']", keys=person.right_diastolic_pressure)


def enter_right_systolic_pressure(driver: WebDriver, person: Person):
    send_keys_by_xpath(driver, path="//input[@name='constriction']", keys=person.right_systolic_pressure)


def enter_breath_rate(driver: WebDriver, person: Person):
    send_keys_by_xpath(driver, path="//input[@name='breathe']", keys=person.breath_rate)


def enter_pulse_rate(driver: WebDriver, person: Person):
    send_keys_by_xpath(driver, path="//input[@name='pulse']", keys=person.pulse_rate)


def enter_temperature(driver: WebDriver, person: Person):
    send_keys_by_xpath(driver, path="//input[@name='temperature']", keys=person.temperature)


def send_keys_by_xpath(driver: WebDriver, path: str, keys: str):
    send_keys_if_keys_not_empty(element=find_element_by_xpath(driver, path), keys=keys)


def send_keys_if_keys_not_empty(element: WebElement, keys: str):
    if keys is None or len(keys.strip()) == 0:
        return
    element.clear()
    element.send_keys(keys)


def import_symptom(driver: WebDriver, person: Person):
    select_no_symptom(driver, person)


def import_check_way(driver: WebDriver, person: Person):
    select_elder(driver, person)
    select_hypertensive(driver, person)
    select_diabetic(driver, person)


def select_no_symptom(driver: WebDriver, person: Person):
    click_select_by_xpath(driver=driver, path="//input[@name='symptom'][@value='01']")


def select_diabetic(driver: WebDriver, person: Person):
    if person.is_diabetic():
        click_select_by_xpath(driver=driver, path="//input[@name='checkWay'][@value=4]")


def select_hypertensive(driver: WebDriver, person: Person):
    if person.is_hypertensive():
        click_select_by_xpath(driver=driver, path="//input[@name='checkWay'][@value=3]")


def select_elder(driver: WebDriver, person: Person):
    if person.is_elder():
        click_select_by_xpath(driver=driver, path="//input[@name='checkWay'][@value=2]")


def click_select_by_xpath(driver: WebDriver, path: str):
    click_select_if_no_selected(find_element_by_xpath(driver, path))


def click_select_if_no_selected(element: WebElement):
    if not element.is_selected():
        element.click()


def click_create_button(driver: WebDriver):
    time.sleep(0.5)
    find_element_by_xpath(driver, "//button[text()='新建(F2)']").click()


def double_click_result_list(driver: WebDriver):
    xpath = "/html/body/div[1]/div/div/div[2]/table/tbody/tr[1]/td[3]/div/div[2]/div/div[1]/div[2]/div[1]/div/div/div/div[2]/div/div/div/div/div[1]/div[1]/div/div/div/div[2]/div[1]/div/div/div[2]/div/div[1]/div[2]/div/div"
    element: WebElement = WebDriverWait(driver, 10).until(
        expected_conditions.element_to_be_clickable((By.XPATH, xpath)))
    ActionChains(driver).double_click(element).perform()


def click_search_button(driver: WebDriver):
    find_search_button(driver).click()


def find_search_button(driver: WebDriver) -> WebElement:
    search_button_xpath_list = [
        "/html/body/div[1]/div/div/div[2]/table/tbody/tr[1]/td[3]/div/div[2]/div/div/div[2]/div[1]/div/div/div/div[2]/div/div/div/div/div[1]/div[1]/div/div/div/div[2]/div[1]/div/div/div[1]/div/table/tbody/tr/td[1]/table/tbody/tr/td[5]/table/tbody/tr[2]/td[2]/em/button",
        "/html/body/div[1]/div/div/div[2]/table/tbody/tr[1]/td[3]/div/div[2]/div/div[1]/div[2]/div[1]/div/div/div/div[2]/div/div/div/div/div[1]/div[1]/div/div/div/div[2]/div[1]/div/div/div[1]/div/table/tbody/tr/td[1]/table/tbody/tr/td[5]/table/tbody/tr[2]/td[2]/em/button"
    ]
    for xpath in search_button_xpath_list:
        try:
            return find_element_by_xpath(driver, xpath)
        except Exception as e:
            pass


def enter_person_id(driver: WebDriver, person_id: str):
    id_input: WebElement = find_id_input(driver)
    id_input.clear()
    id_input.send_keys(person_id)


def find_id_input(driver: WebDriver) -> WebElement:
    id_input_xpath_list = [
        "/html/body/div[1]/div/div/div[2]/table/tbody/tr[1]/td[3]/div/div[2]/div/div/div[2]/div[1]/div/div/div/div[2]/div/div/div/div/div[1]/div[1]/div/div/div/div[2]/div[1]/div/div/div[1]/div/table/tbody/tr/td[1]/table/tbody/tr/td[3]/input",
        "/html/body/div[1]/div/div/div[2]/table/tbody/tr[1]/td[3]/div/div[2]/div/div[1]/div[2]/div[1]/div/div/div/div[2]/div/div/div/div/div[1]/div[1]/div/div/div/div[2]/div[1]/div/div/div[1]/div/table/tbody/tr/td[1]/table/tbody/tr/td[3]/input"
    ]
    for xpath in id_input_xpath_list:
        try:
            return find_element_by_xpath(driver, xpath)
        except Exception as e:
            pass


def click_id_limit_select(driver: WebDriver):
    time.sleep(0.5)
    find_id_limit_select(driver).click()


def find_id_limit_select(driver: WebDriver) -> WebElement:
    id_limit_xpath_list: List[str] = [
        "/html/body/div[12]/div/div[6]",
        "/html/body/div[8]/div/div[6]",
        "/html/body/div[17]/div/div[6]"
    ]
    for xpath in id_limit_xpath_list:
        try:
            return find_element_by_xpath(driver, xpath)
        except Exception as e:
            pass


def click_limit_select_img(driver: WebDriver):
    time.sleep(0.5)
    find_limit_select_input(driver).click()


def find_limit_select_input(driver: WebDriver) -> WebElement:
    input_xpath_list: List[str] = [
        "/html/body/div[1]/div/div/div[2]/table/tbody/tr[1]/td[3]/div/div[2]/div/div[1]/div[2]/div[1]/div/div/div/div[2]/div/div/div/div/div[1]/div[1]/div/div/div/div[2]/div[1]/div/div/div[1]/div/table/tbody/tr/td[1]/table/tbody/tr/td[1]/div/input",
        "/html/body/div[1]/div/div/div[2]/table/tbody/tr[1]/td[3]/div/div[2]/div/div/div[2]/div[1]/div/div/div/div[2]/div/div/div/div/div[1]/div[1]/div/div/div/div[2]/div[1]/div/div/div[1]/div/table/tbody/tr/td[1]/table/tbody/tr/td[1]/div/input"]
    for input_xpath in input_xpath_list:
        try:
            return find_element_by_xpath(driver, input_xpath)
        except Exception as e:
            pass


def click_health_check_sheet_li(driver: WebDriver):
    time.sleep(0.5)
    find_element_by_id(driver, "WL_module_D20").click()


def click_health_doc_li(driver: WebDriver):
    WebDriverWait(driver, 10).until(expected_conditions.element_to_be_clickable((By.ID, "HR")))
    find_element_by_id(driver, "HR").click()


def click_login_button(driver: WebDriver):
    time.sleep(0.5)
    find_element_by_id(driver, "logon").click()


def select_role(driver: WebDriver):
    time.sleep(0.5)
    role_input: WebElement = find_element_by_xpath(driver, "//div[@id='select-role']//div//input")
    role_input.click()
    time.sleep(0.5)
    find_element_by_xpath(driver, "//li[text()='责任医生']//parent::ul").click()
    time.sleep(0.5)
    find_element_by_xpath(driver, "//li[text()='责任医生']//parent::ul").click()


def set_username(driver: WebDriver, username: str):
    driver.find_element(by=By.XPATH, value="//div[@id='usertext']//div//input").send_keys(username)


def set_password(driver: WebDriver, password: str):
    driver.find_element(by=By.ID, value="pwd").send_keys(password)


def import_cover(username: str, password: str, file_path: str):
    print(username, password)


def find_element_by_xpath(driver: WebDriver, path: str) -> WebElement:
    return driver.find_element(by=By.XPATH, value=path)


def find_element_by_id(driver: WebDriver, html_id: str) -> WebElement:
    return driver.find_element(by=By.ID, value=html_id)


def read_excel(file_path: str) -> List[Person]:
    workbook: Workbook = load_workbook(file_path)
    sheet: Worksheet = workbook.active
    rows: Generator[tuple[Cell, ...], None, None] = sheet.iter_rows(min_row=config.current_row_index)
    persons: List[Person] = []
    for row in rows:
        person = Person(row)
        persons.append(person)
    return persons
